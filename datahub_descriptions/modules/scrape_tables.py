# Import modules
from openpyxl import load_workbook
import tabula
import pandas as pd
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class TableScraperDC:
    """ Class to scrape pdfs for tables."""
    
    def __init__(self, pdf_path, xlsx_path):
        self.pdf_path = pdf_path
        self.xlsx_path = xlsx_path
        logging.info("TableScraperDC initialized")

    def extract_tables_from_pdf(self, pdf_path: str) -> list[pd.DataFrame]:
        """
        Function to extract the tables from the pdf using tabula.

        Input:
        - pdf_path = path to pdf, in this case UCube structure.pdf

        Output:
        - list of pandas DataFrames containing all tables in the pdf
        """
        try:
            logging.info("Starting table extraction from PDF: %s", pdf_path)
            tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
            logging.info("Table extraction successful. Number of tables extracted: %d", len(tables))
            return tables
        
        except Exception as e:
            logging.error("Error during table extraction: %s", e)
            return []


    def ucube_format_pdf(self, tables: list[pd.DataFrame]) -> dict[str, pd.DataFrame]:
        """ 
        Function to scrape the UCube structure.pdf 
        
        Input:
        - tables : list of DataFrames extracted using extract_tables()

        Output:
        - dictionary of tables in the format: {"table_name": pandas_df}

        The dataframes output have 2 columns - column_name and description.
        """
        try:
            logging.info("Starting UCube formatting for pdf format.")
            current_table_name = None
            tables_dict = {}
            
            # Tables need to be merged correctly since tabula splits them if they are on different pages
            for table in tables:
                # Take only description columns of tables
                df = pd.DataFrame(table)
                table_name = df.columns[0]
                df = df.iloc[:, :2]
                df.columns = ['column_name', 'description']

                if table_name == current_table_name:
                    # This is for tables that need to be merged over multiple tables
                    df_prev = tables_dict[current_table_name]
                    df_combined = pd.concat([df_prev, df], axis=0, ignore_index=True)
                    tables_dict[table_name] = df_combined
                    current_table_name = table_name
                else:
                    # If tables don't already exist, then make a new key in dictionary
                    tables_dict[table_name] = df
                    current_table_name = table_name

            # Drop last 2 keys - they are tables we don't need
            tables_dict_final = {key: value for key, value in tables_dict.items() if key not in list(tables_dict.keys())[-2:]}
            
            logging.info("UCube formatting completed successfully.")
            return tables_dict_final
        
        except Exception as e:
            logging.error("Error during UCube formatting: %s", e)
            return {}
        
    
    def extract_tables_from_xlsx(self, xlsx_path:str) -> dict[str, pd.DataFrame]:
        """
        Function to extract the tables from the xlsx using openpyxl. 
        Formatting is done in the same step as we are using sheets in the XLSX, so this step merges ucube_format_pdf()

        Input:
        - xlsx_path = path to xlsx, in this case UCube structure.xlsx

        Output:
        - list of pandas DataFrames containing all tables in the xlsx file
        """
        try:
            logging.info("Starting table extraction from XLSX: %s", xlsx_path)
            wb = load_workbook(filename=xlsx_path, data_only=True)
            sheet = wb['Table attributes']

            tables_dict = {}
            start_capturing = False
            current_table = None
            data = []

            for row in sheet.iter_rows(values_only=True):
                # Start capturing data when the 'Asset' table is found
                if row[0] == 'Asset':
                    start_capturing = True
                    logging.info("Starting to capture tables...")
                if start_capturing:
                    # New table name found
                    if row[0] and all(cell is None for cell in row[1:]):
                        # If there's already data from a previous table, create a DataFrame for it
                        if current_table and data:
                            df = pd.DataFrame(data, columns=['column_name', 'description'])
                            tables_dict[current_table] = df
                            logging.info(f"Finished processing table: {current_table}")
                            data = []  # Reset data for the next table
                        current_table = row[0]
                        logging.info(f"Processing new table: {current_table}")
                    # Get column name and description
                    elif current_table and row[0] and row[1]:
                        data.append([row[0], row[1]])

            # Adding the last table after the loop ends
            if current_table and data:
                df = pd.DataFrame(data, columns=['column_name', 'description'])
                tables_dict[current_table] = df
                logging.info(f"Finished processing table: {current_table}")
                        
            return tables_dict
        
        except Exception as e:
            logging.error("Error during table extraction: %s", e)
            return []
        